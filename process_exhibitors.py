import dataclasses
from typing import List, Set, Tuple, Union
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

base = "https://jaarbeurszakelijk.app.swapcard.com"

class_company_name = "dbnofQ"
class_company_info = "gbyhbR"
class_segment = "gShnkr"
class_contact = "btWXRc"
class_social_media = "kzrhIj"


@dataclasses.dataclass
class Company:
    name: str = ''
    address: str = ''
    country: str = ''
    segments: Set[str] = dataclasses.field(default_factory=set)
    social_media: List[str] = dataclasses.field(default_factory=list)
    websites: List[str] = dataclasses.field(default_factory=list)
    emails: List[str] = dataclasses.field(default_factory=list)
    telephones: List[str] = dataclasses.field(default_factory=list)
    information: List[str] = dataclasses.field(default_factory=list)

    def init_from_soup(self, soup: BeautifulSoup):
        self.find_info(soup)
        self.find_contact_details(soup)

    def find_info(self, soup):
        name_element = soup.find(class_=class_company_name)
        if name_element:
            self.name = name_element.text

        info_element = soup.find(class_=class_company_info)
        if info_element:
            paragraphs = info_element.find_all('p')
            for p in paragraphs:
                self.information.append(p.text)

    def find_contact_details(self, soup: BeautifulSoup):
        contact_elements = soup.findAll(class_=class_contact)
        if contact_elements:
            for contact_element in contact_elements:
                contact_element_text = contact_element.text
                if '@' in contact_element_text:
                    self.emails.append(contact_element_text)
                elif 'https' in contact_element_text:
                    self.websites.append(contact_element_text)
                elif '+' in contact_element_text:
                    self.telephones.append(contact_element_text)
                elif ',' in contact_element_text:
                    self.country = contact_element_text.split(',')[-1]
                    self.address = contact_element_text
                else:
                    pass


with open("exhibitor_urls.txt") as f:
    wb = Workbook()
    ws = wb.active
    ws.title = 'cyberseceurope'
    ws['A1'] = 'name'
    ws['B1'] = 'commentaar'
    ws['C1'] = 'country'
    ws['D1'] = 'address'
    ws['E1'] = 'websites'
    ws['F1'] = 'emails'
    ws['G1'] = 'telephones'
    ws['H1'] = 'information'
    count = 2
    for widget in f.readlines():
        url = base + widget.strip()
        print(url)
        response = requests.get(url)
        if response.status_code == 200:
            contents = response.text
            soup = BeautifulSoup(contents, "html.parser")
            company = Company()
            company.find_info(soup)
            company.find_contact_details(soup)

            ws[f'A{count}'] = company.name
            ws[f'B{count}'] = company.country
            ws[f'C{count}'] = company.address
            if company.websites:
                ws[f'D{count}'] = company.websites[0]
            if company.emails:
                ws[f'E{count}'] = company.emails[0]
            if company.telephones:
                ws[f'F{count}'] = company.telephones[0]
            if company.information:
                ws[f'G{count}'] = company.information[0]

            count += 1

    wb.save('cyberseceurope_bedrijven.xlsx')
